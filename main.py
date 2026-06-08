import streamlit as st
import pandas as pd
import math
import io
import zipfile
from openpyxl import load_workbook
from datetime import datetime

st.set_page_config(page_title="Batch File Splitter", layout="centered")

st.title("📂 Excel / CSV Batch Splitter")
st.write("Upload a file and split it into multiple batches based on number of rows.")

uploaded_file = st.file_uploader(
    "Upload Excel / CSV file",
    type=["csv", "xlsx", "xls"]
)

rows_per_batch = st.number_input(
    "Enter rows per batch",
    min_value=1,
    step=1
)

if uploaded_file is not None and rows_per_batch:

    file_name = uploaded_file.name.lower()

    try:
        if file_name.endswith(".csv"):
            df = pd.read_csv(uploaded_file)

        elif file_name.endswith(".xlsx"):
            df = pd.read_excel(uploaded_file, engine="openpyxl")

        elif file_name.endswith(".xls"):
            df = pd.read_excel(uploaded_file, engine="xlrd")

        else:
            st.error("❌ Unsupported file format")
            st.stop()

    except Exception as e:
        st.error(f"❌ Error reading file: {e}")
        st.stop()

    # Auto-detect date columns
    for col in df.columns:
        try:
            converted = pd.to_datetime(df[col], errors="coerce")

            # If at least 80% values are valid dates,
            # treat the entire column as a date column
            if converted.notna().sum() > len(df) * 0.8:
                df[col] = converted

        except Exception:
            pass

    total_rows = len(df)
    total_batches = math.ceil(total_rows / rows_per_batch)

    st.success(f"✅ Total Rows: {total_rows}")
    st.info(f"📦 Total Batches: {total_batches}")

    if st.button("🚀 Split File"):

        progress = st.progress(0)

        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:

            for i in range(total_batches):

                start = i * rows_per_batch
                end = start + rows_per_batch

                batch_df = df.iloc[start:end]

                temp_output = io.BytesIO()

                # Write batch to Excel
                with pd.ExcelWriter(temp_output, engine="openpyxl") as writer:
                    batch_df.to_excel(
                        writer,
                        index=False,
                        sheet_name="Sheet1"
                    )

                temp_output.seek(0)

                # Re-open workbook for formatting
                wb = load_workbook(temp_output)
                ws = wb.active

                # Loop through all cells
                for row in ws.iter_rows(min_row=2):
                    for cell in row:

                        if isinstance(
                            cell.value,
                            (datetime, pd.Timestamp)
                        ):
                            cell.number_format = "DD/MMM/YYYY"

                final_output = io.BytesIO()
                wb.save(final_output)

                batch_filename = f"batch_{i + 1}.xlsx"

                zf.writestr(
                    batch_filename,
                    final_output.getvalue()
                )

                progress.progress((i + 1) / total_batches)

        st.success("✅ File successfully split into batches")

        st.download_button(
            label="📥 Download ZIP",
            data=zip_buffer.getvalue(),
            file_name="split_batches.zip",
            mime="application/zip"
        )
